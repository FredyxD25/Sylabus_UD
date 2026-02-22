import os
import sys
import glob
import re
from openpyxl import load_workbook

REEMPLAZOS = {
    "Versión: 01": "Versión: 02",
    "Version: 01": "Version: 02",
}

def procesar_celda(valor, patron_fecha, nuevo_texto_fecha):
    if not isinstance(valor, str):
        return valor, False
    modificado = False
    for buscar, reemplazar in REEMPLAZOS.items():
        if buscar in valor:
            valor = valor.replace(buscar, reemplazar)
            modificado = True
    nuevo_valor, n = re.subn(patron_fecha, nuevo_texto_fecha, valor)
    if n > 0:
        valor = nuevo_valor
        modificado = True
    return valor, modificado

def procesar_excel(ruta_archivo, patron_fecha, nuevo_texto_fecha):
    try:
        wb = load_workbook(ruta_archivo)
        archivo_modificado = False
        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            for fila in hoja.iter_rows():
                for celda in fila:
                    nuevo_valor, modificado = procesar_celda(celda.value, patron_fecha, nuevo_texto_fecha)
                    if modificado:
                        celda.value = nuevo_valor
                        archivo_modificado = True
        if archivo_modificado:
            wb.save(ruta_archivo)
            print(f"✅ Actualizado: {os.path.basename(ruta_archivo)}")
        else:
            print(f"⏭️  Sin cambios: {os.path.basename(ruta_archivo)}")
    except Exception as e:
        print(f"❌ Error: {os.path.basename(ruta_archivo)} → {e}")

def main(carpeta, nueva_fecha):
    patron_fecha = r"Fecha de Aprobación:\s*\d{1,2}/\d{1,2}/\d{4}"
    nuevo_texto_fecha = f"Fecha de Aprobación:\n{nueva_fecha}"

    excels = glob.glob(os.path.join(carpeta, "**", "*.xlsx"), recursive=True)
    excels += glob.glob(os.path.join(carpeta, "**", "*.xls"), recursive=True)

    if not excels:
        print("⚠️  No se encontraron archivos Excel.")
        return

    print(f"📂 Carpeta: {carpeta}")
    print(f"📊 Excel encontrados: {len(excels)}")
    print(f"🔄 Versión: 01 → 02")
    print(f"📅 Nueva fecha: {nueva_fecha}")
    print("=" * 55)

    for archivo in excels:
        procesar_excel(archivo, patron_fecha, nuevo_texto_fecha)

    print("=" * 55)
    print("✔️  Proceso completado.")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("❌ Uso: python ActualizarCampos.py <carpeta> <nueva_fecha>")
        print("   Ejemplo: python ActualizarCampos.py C:\\mis\\excels 15/03/2024")
        sys.exit(1)

    carpeta   = sys.argv[1]
    nueva_fecha = sys.argv[2]

    if not os.path.isdir(carpeta):
        print(f"❌ La carpeta no existe: {carpeta}")
        sys.exit(1)

    main(carpeta, nueva_fecha)