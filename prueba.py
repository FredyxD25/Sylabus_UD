import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.cell.cell import Cell
from copy import copy
import os

# Ruta de entrada y salida
archivo_entrada = "archivo1.xlsx"
archivo_salida = "estructura_vacia.xlsx"

# Cargar archivo original
wb_origen = openpyxl.load_workbook(archivo_entrada)
wb_nuevo = openpyxl.Workbook()
wb_nuevo.remove(wb_nuevo.active)  # Eliminar hoja por defecto

for hoja_nombre in wb_origen.sheetnames:
    hoja_origen = wb_origen[hoja_nombre]
    hoja_nueva = wb_nuevo.create_sheet(title=hoja_nombre)

    # Copiar dimensiones de columnas y filas
    for col in hoja_origen.column_dimensions:
        hoja_nueva.column_dimensions[col].width = hoja_origen.column_dimensions[col].width
    for row in hoja_origen.row_dimensions:
        hoja_nueva.row_dimensions[row].height = hoja_origen.row_dimensions[row].height

    # Copiar celdas (sin valores) pero con estilo
    for fila in hoja_origen.iter_rows():
        for celda in fila:
            if isinstance(celda, Cell):  # Ignorar MergedCell
                nueva_celda = hoja_nueva.cell(row=celda.row, column=celda.col_idx)
                nueva_celda.font = copy(celda.font)
                nueva_celda.fill = copy(celda.fill)
                nueva_celda.border = copy(celda.border)
                nueva_celda.alignment = copy(celda.alignment)
                nueva_celda.number_format = copy(celda.number_format)

    # Copiar celdas combinadas
    for rango in hoja_origen.merged_cells.ranges:
        hoja_nueva.merge_cells(str(rango))

    # Copiar imágenes si están disponibles (solo si tienen un archivo de origen)
    if hasattr(hoja_origen, '_images'):
        for imagen in hoja_origen._images:
            try:
                nueva_imagen = OpenpyxlImage(imagen.ref)  # Esto solo funciona si hay ruta de archivo
                nueva_imagen.anchor = imagen.anchor
                hoja_nueva.add_image(nueva_imagen)
            except Exception as e:
                print(f"No se pudo copiar una imagen: {e}")

# Guardar archivo
wb_nuevo.save(archivo_salida)
print(f"Archivo exportado: {archivo_salida}")
